#!/usr/bin/env python3
"""
Cyber5 Chess Live Tournament Display
A Flask-based web application for real-time chess tournament display
with Excel file monitoring via WebSockets.
"""

import os
import json
import threading
import logging
from pathlib import Path
from datetime import datetime
from flask import Flask, render_template, request, jsonify
from flask_socketio import SocketIO, emit, join_room, leave_room
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
import pandas as pd
from openpyxl import load_workbook

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Initialize Flask app
app = Flask(__name__)
app.config['SECRET_KEY'] = 'cyber5-chess-tournament-secret-key-2026'
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading')

# Configuration
WORKBOOK_PATH = os.path.expanduser(
    '~/Library/CloudStorage/OneDrive-AmericanExp/Cyber5 Indoor Sports Tournament 2026 – Chess Tournament.xlsx'
)
DEFAULT_SHEET_NAME = 'Match Tracker'
POLL_INTERVAL = 2  # seconds

# Global state
class TournamentState:
    def __init__(self):
        self.matches = []
        self.last_modified = None
        self.lock = threading.Lock()
        self.file_observer = None
        self.monitor_thread = None
        self.is_running = True

state = TournamentState()


class ExcelFileEventHandler(FileSystemEventHandler):
    """Handles file system events for Excel file changes."""
    
    def on_modified(self, event):
        if not event.is_directory and event.src_path.endswith('.xlsx'):
            if WORKBOOK_PATH in event.src_path:
                logger.info(f"Excel file modified: {event.src_path}")
                threading.Thread(target=reload_tournament_data, daemon=True).start()


def read_excel_matches(file_path: str) -> list:
    """
    Read match data from Excel workbook.
    
    Args:
        file_path: Path to the Excel workbook
        
    Returns:
        List of match dictionaries
    """
    try:
        if not os.path.exists(file_path):
            logger.warning(f"Excel file not found: {file_path}")
            return []
        
        # Read the workbook
        wb = load_workbook(file_path, data_only=True)
        
        # Try to get the Match Tracker sheet
        if DEFAULT_SHEET_NAME not in wb.sheetnames:
            logger.warning(f"Sheet '{DEFAULT_SHEET_NAME}' not found. Available sheets: {wb.sheetnames}")
            return []
        
        ws = wb[DEFAULT_SHEET_NAME]
        matches = []
        
        # Skip header row (row 1)
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Skip empty rows
            if not any(row):
                continue
            
            # Map Excel columns to match data
            # Expected columns: Match ID, Player 1, Player 2, Winner, Status, Round, Result
            match_data = {
                'id': row[0] if row[0] else '',
                'player1': row[1] if len(row) > 1 and row[1] else 'TBD',
                'player2': row[2] if len(row) > 2 and row[2] else 'TBD',
                'winner': row[3] if len(row) > 3 and row[3] else '',
                'status': row[4] if len(row) > 4 and row[4] else 'Pending',
                'round': row[5] if len(row) > 5 and row[5] else '',
                'result': row[6] if len(row) > 6 and row[6] else '',
            }
            
            # Only add matches with valid ID
            if match_data['id']:
                matches.append(match_data)
        
        wb.close()
        logger.info(f"Successfully read {len(matches)} matches from Excel")
        return matches
        
    except PermissionError:
        logger.error(f"Permission denied accessing Excel file: {file_path}")
        return []
    except Exception as e:
        logger.error(f"Error reading Excel file: {str(e)}")
        return []


def reload_tournament_data():
    """
    Reload tournament data from Excel and broadcast updates via WebSocket.
    """
    try:
        with state.lock:
            new_matches = read_excel_matches(WORKBOOK_PATH)
            
            # Check if data changed
            if new_matches != state.matches:
                state.matches = new_matches
                state.last_modified = datetime.now().isoformat()
                
                # Broadcast update to all connected clients
                socketio.emit('matches_updated', {
                    'matches': state.matches,
                    'timestamp': state.last_modified
                }, broadcast=True)
                
                logger.info(f"Tournament data updated. Broadcasting to all clients.")
    except Exception as e:
        logger.error(f"Error reloading tournament data: {str(e)}")


def monitor_excel_file():
    """
    Background thread that monitors Excel file for changes.
    """
    logger.info("Starting Excel file monitor...")
    
    # Initial load
    reload_tournament_data()
    
    # Set up file system observer
    event_handler = ExcelFileEventHandler()
    observer = Observer()
    
    # Get directory of the workbook
    workbook_dir = os.path.dirname(os.path.expanduser(WORKBOOK_PATH))
    
    try:
        observer.schedule(event_handler, workbook_dir, recursive=False)
        observer.start()
        logger.info(f"Monitoring directory: {workbook_dir}")
        
        # Keep polling as fallback
        while state.is_running:
            if state.last_modified is None or \
               (datetime.now() - datetime.fromisoformat(state.last_modified)).total_seconds() > POLL_INTERVAL:
                reload_tournament_data()
            
            threading.Event().wait(POLL_INTERVAL)
            
    except Exception as e:
        logger.error(f"Error in file monitor: {str(e)}")
    finally:
        observer.stop()
        observer.join()


# Routes
@app.route('/')
def index():
    """
    Render the main tournament display page.
    """
    return render_template('index.html')


@app.route('/api/matches')
def get_matches():
    """
    API endpoint to get current matches.
    """
    with state.lock:
        return jsonify({
            'matches': state.matches,
            'timestamp': state.last_modified,
            'workbook_path': WORKBOOK_PATH
        })


@app.route('/api/health')
def health():
    """
    Health check endpoint.
    """
    return jsonify({
        'status': 'healthy',
        'matches_count': len(state.matches),
        'workbook_exists': os.path.exists(WORKBOOK_PATH),
        'timestamp': datetime.now().isoformat()
    })


# WebSocket Events
@socketio.on('connect')
def handle_connect():
    """
    Handle client connection.
    """
    logger.info(f"Client connected: {request.sid}")
    join_room('tournament_room')
    
    # Send initial data
    with state.lock:
        emit('initial_data', {
            'matches': state.matches,
            'timestamp': state.last_modified
        })


@socketio.on('disconnect')
def handle_disconnect():
    """
    Handle client disconnection.
    """
    logger.info(f"Client disconnected: {request.sid}")
    leave_room('tournament_room')


@socketio.on('request_refresh')
def handle_refresh_request():
    """
    Handle manual refresh request from client.
    """
    logger.info("Manual refresh requested")
    reload_tournament_data()


if __name__ == '__main__':
    # Start Excel file monitor thread
    monitor_thread = threading.Thread(target=monitor_excel_file, daemon=True)
    monitor_thread.start()
    logger.info("Monitor thread started")
    
    # Start Flask-SocketIO server
    logger.info("Starting Cyber5 Chess Tournament Display on http://localhost:5000")
    socketio.run(
        app,
        host='127.0.0.1',
        port=5000,
        debug=True,
        use_reloader=False,
        log_output=True
    )
